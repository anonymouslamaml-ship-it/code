import torch
import numpy as np
import random
from functools import partial
from environment import (GoToLocalMissionEnv,
                         GoToOpenMissionEnv, 
                         GoToObjDoorMissionEnv,  
                         PickupDistMissionEnv,
                         OpenDoorMissionEnv, 
                         OpenDoorLocMissionEnv,
                         OpenDoorsOrderMissionEnv)
from sampler_lang import BabyAIMissionTaskWrapper, SentenceMissionEncoder, MissionParamAdapter
import sampler_lang 
import sampler_maml
import sampler_lang_policy
import sampler_anil
import os
from maml_rl.policies.categorical_mlp import CategoricalMLPPolicy
from maml_rl.utils.reinforcement_learning import reinforce_loss
from maml_rl.episode import BatchEpisodes
from maml_rl.baseline import LinearFeatureBaseline
from collections import OrderedDict
import argparse
from openpyxl import Workbook, load_workbook
import builtins, io
from contextlib import contextmanager, redirect_stdout, redirect_stderr

@contextmanager
def silence_sampling_rejected():
    real_print = builtins.print
    buf = io.StringIO()
    def filtered_print(*args, **kwargs):
        if args and isinstance(args[0], str) and args[0].startswith("Sampling rejected: unreachable object"):
            return
        return real_print(*args, **kwargs)
    builtins.print = filtered_print
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            yield
    finally:
        builtins.print = real_print


seed = 42
random.seed(seed)
np.random.seed(seed)
torch.manual_seed(seed)
if torch.cuda.is_available():
    torch.cuda.manual_seed_all(seed)

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

p = argparse.ArgumentParser()
p.add_argument("--env", dest="env_name",
               choices=["GoToLocal","PickupDist","GoToObjDoor","GoToOpen","OpenDoor",
                        "OpenDoorLoc","OpenDoorsOrder"],
               default="GoToLocal")
p.add_argument("--room-size", type=int, default=7)
p.add_argument("--num-dists", type=int, default=3)
p.add_argument("--max-steps", type=int, default=300)
p.add_argument("--delta-theta", type=float, default=0.3)
p.add_argument("--skip-random", action="store_true",
               help="Skip the random-policy baseline to speed up evaluation")

args = p.parse_args()


OBJECTS = ['box']
COLORS = ['red', 'green', 'blue', 'purple','yellow', 'grey']
PREP_LOCS = ['on', 'at', 'to']

# Location names
LOC_NAMES = ['right', 'front']

DOOR_COLORS = ['yellow', 'grey']

# For Pickup
PICKUP_MISSIONS = [f"pick up the {color} {obj}" for color in COLORS for obj in OBJECTS]

# For GoToLocal
LOCAL_MISSIONS = [f"go to the {color} {obj}" for color in COLORS for obj in OBJECTS]

# For environments that include doors (GoToObjDoor, GoToOpen, Open)
DOOR_MISSIONS = [f"go to the {color} door" for color in DOOR_COLORS]
OPEN_DOOR_MISSIONS = [f"open the {color} door" for color in DOOR_COLORS]
DOOR_LOC_MISSIONS = [f"open the door {prep} the {loc}" for prep in PREP_LOCS for loc in LOC_NAMES]
OPEN_DOORS_ORDER_MISSIONS = (
    [f"open the {c1} door" for c1 in DOOR_COLORS] +
    [f"open the {c1} door, then open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS] +
    [f"open the {c1} door after you open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS]
)


def build_env(env, room_size, num_dists, max_steps, missions):

    if env == "GoToLocal":
        base = GoToLocalMissionEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps)
    elif env == "PickupDist":
        base = PickupDistMissionEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps)
    elif env == "GoToObjDoor":
        base = GoToObjDoorMissionEnv(max_steps=max_steps, num_distractors=num_dists)
    elif env == "GoToOpen":
        base = GoToOpenMissionEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps)
    elif env == "OpenDoor":
        base = OpenDoorMissionEnv(room_size=room_size, max_steps=max_steps)
    elif env == "OpenDoorLoc":
        base = OpenDoorLocMissionEnv(room_size=room_size, max_steps=max_steps)
    elif env == "OpenDoorsOrder":
        base = OpenDoorsOrderMissionEnv(room_size=room_size)
    else:
        raise ValueError(f"Unknown env_name: {env}")

    return BabyAIMissionTaskWrapper(base, missions=missions)


def select_missions(env_name):
    mission_map = {
        "GoToLocal": LOCAL_MISSIONS,
        "PickupDist": PICKUP_MISSIONS,
        "GoToObjDoor": LOCAL_MISSIONS + DOOR_MISSIONS,
        "GoToOpen": LOCAL_MISSIONS,
        "OpenDoor": OPEN_DOOR_MISSIONS,
        "OpenDoorLoc": OPEN_DOOR_MISSIONS + DOOR_LOC_MISSIONS,
        "OpenDoorsOrder": OPEN_DOORS_ORDER_MISSIONS    }
    return mission_map[env_name]


env_name  = args.env_name
room_size = "env" if args.env_name in ["GoToObjDoor"] else args.room_size
num_dists = "env" if args.env_name in ["OpenDoor", "OpenDoorLoc", "OpenDoorsOrder"] else args.num_dists
max_steps = args.max_steps
delta_theta = args.delta_theta

missions = select_missions(env_name)
make_env = partial(build_env, env_name, room_size, num_dists, max_steps, missions)
env = make_env()

print(f"env name {env} \n")
print(f"room_size: {room_size}\nnum_dists: {num_dists}\nmax_steps: {max_steps}\n")


# ==================== SETUP: LA-MAML Policy ====================
ckpt = torch.load(f"lang_model/lang_{env_name}_{delta_theta}.pth", map_location=device)

dummy_obs, _ = env.reset()
input_size_lang = sampler_lang.preprocess_obs(dummy_obs).shape[0]
output_size = env.action_space.n
hidden_sizes = (64, 64)
nonlinearity = torch.nn.functional.tanh

# Policy language
policy_lang = CategoricalMLPPolicy(
    input_size=input_size_lang,
    output_size=output_size,
    hidden_sizes=hidden_sizes,
    nonlinearity=nonlinearity,
).to(device)  
policy_lang.load_state_dict(ckpt["policy"])
policy_lang.eval()
policy_param_shapes = [p.shape for p in policy_lang.parameters()]

mission_encoder = SentenceMissionEncoder(
    model_name="all-MiniLM-L6-v2",
    frozen=True,
    normalize=True,
    cache=True,
    device=device
)
mission_encoder.eval()

preprocess_obs = sampler_lang.preprocess_obs
mission_encoder_output_dim = mission_encoder.output_dim

# Adapter
mission_adapter = MissionParamAdapter(mission_encoder_output_dim, policy_param_shapes).to(device)
mission_adapter.load_state_dict(ckpt["mission_adapter"])    
mission_adapter.eval()


# ==================== SETUP: MAML Policy ====================
ckpt_base_list = [f"maml_model/maml_{env_name}.pth", f"maml_model/maml_{env_name}_1.pth"]
ckpt_base = next((p for p in ckpt_base_list if os.path.exists(p)), ckpt_base_list[0])
ckpt_maml = torch.load(ckpt_base, map_location=device)

dummy_obs, _ = env.reset()
input_size_maml = sampler_maml.preprocess_obs(dummy_obs).shape[0]

policy_maml = CategoricalMLPPolicy(
    input_size=input_size_maml,
    output_size=output_size,
    hidden_sizes=hidden_sizes,
    nonlinearity=nonlinearity,
).to(device)

if isinstance(ckpt_maml, dict) and "policy" in ckpt_maml:
    policy_maml.load_state_dict(ckpt_maml["policy"])
else:
    policy_maml.load_state_dict(ckpt_maml)

policy_maml.eval()

baseline = LinearFeatureBaseline(input_size_maml).to(device)


# ==================== SETUP: ANIL Policy ====================
ckpt_anil_list = [f"anil_model/anil_{env_name}.pth", f"anil_model/anil_{env_name}_1.pth"]
ckpt_anil_path = next((p for p in ckpt_anil_list if os.path.exists(p)), ckpt_anil_list[0])
ckpt_anil = torch.load(ckpt_anil_path, map_location=device)

input_size_anil = sampler_anil.preprocess_obs(dummy_obs).shape[0]

policy_anil = CategoricalMLPPolicy(
    input_size=input_size_anil,
    output_size=output_size,
    hidden_sizes=hidden_sizes,
    nonlinearity=nonlinearity,
).to(device)

if isinstance(ckpt_anil, dict) and "policy" in ckpt_anil:
    policy_anil.load_state_dict(ckpt_anil["policy"])
else:
    policy_anil.load_state_dict(ckpt_anil)

policy_anil.eval()



def get_head_layer_prefix(hidden_sizes):
    return f'layer{len(hidden_sizes) + 1}'
head_layer_prefix = get_head_layer_prefix(hidden_sizes)

def update_head_only_eval(policy, loss, params, step_size, head_layer_prefix):
    head_params = OrderedDict((name, param) for name, param in params.items() if head_layer_prefix in name)
    grads = torch.autograd.grad(loss, head_params.values(), create_graph=False)
    updated_params = OrderedDict()
    grad_iter = iter(grads)
    for name, param in params.items():
        updated_params[name] = param - step_size * next(grad_iter) if head_layer_prefix in name else param
    return updated_params

def adapt_anil_policy_for_task(task, policy, env, baseline, num_steps=1, fast_lr=1e-4, batch_size=10, head_layer_prefix='layer3'):
    env.reset_task(task)
    train_batches = []
    for _ in range(num_steps + 1):
        batch = BatchEpisodes(batch_size=batch_size, gamma=0.99, device=device)
        for ep in range(batch_size):
            with silence_sampling_rejected():
                obs, info = env.reset()
            done = False
            episode_obs, episode_actions, episode_rewards = [], [], []
            while not done:
                obs_vec = sampler_anil.preprocess_obs(obs)
                obs_tensor = torch.from_numpy(obs_vec).float().unsqueeze(0).to(device)
                with torch.no_grad():
                    dist = policy(obs_tensor)
                    action = dist.sample().item()
                obs, reward, terminated, truncated, info = env.step(action)
                done = terminated or truncated
                episode_obs.append(obs_vec)
                episode_actions.append(np.array(action))
                episode_rewards.append(np.array(reward, dtype=np.float32))
            batch.append(episode_obs, episode_actions, episode_rewards, [ep]*len(episode_obs))
        train_batches.append(batch)
    for batch in train_batches:
        batch.compute_advantages(baseline, gae_lambda=1.0, normalize=True)
    params = OrderedDict(policy.named_parameters())
    for batch in train_batches:
        loss = reinforce_loss(policy, batch, params=params)
        params = update_head_only_eval(policy, loss, params, fast_lr, head_layer_prefix)
    return params


# ==================== SETUP: Language-conditioned Policy ====================
ckpt_ablation = torch.load(f"lang_policy_model/lang_policy_{env_name}.pth", map_location=device)

sampler_lang_policy.device = device

dummy_obs, _ = env.reset()
dummy_vec = sampler_lang_policy.preprocess_obs(dummy_obs, mission_str=missions[0])
input_size_ablation = dummy_vec.shape[0]

policy_ablation = CategoricalMLPPolicy(
    input_size=input_size_ablation,
    output_size=output_size,
    hidden_sizes=hidden_sizes,
    nonlinearity=nonlinearity,
).to(device)
policy_ablation.load_state_dict(ckpt_ablation["policy"])
policy_ablation.eval()




def get_language_adapted_params(policy, mission_str, mission_encoder, mission_adapter, device):
    with torch.no_grad():
        mission_emb = mission_encoder(mission_str).to(device)  
        delta_thetas = mission_adapter(mission_emb)
        delta_thetas = [delta * delta_theta for delta in delta_thetas]
    policy_params = list(policy.parameters())
    param_names = list(dict(policy.named_parameters()).keys())
    theta_prime = OrderedDict(
        (name, param + delta.squeeze(0))
        for name, param, delta in zip(param_names, policy_params, delta_thetas)
    )
    return theta_prime


def adapt_policy_for_task(task, policy, num_steps=1, fast_lr=1e-4, batch_size=10,baseline=None):
    
    env.reset_task(task)
    
    train_batches = []
    for _ in range(num_steps+1):
        batch = BatchEpisodes(batch_size=batch_size, gamma=0.99, device=device)
        for ep in range(batch_size):
            with silence_sampling_rejected():
                obs, info = env.reset()
            done = False
            episode_obs = []
            episode_actions = []
            episode_rewards = []
            while not done:
                obs_vec = sampler_maml.preprocess_obs(obs)
                obs_tensor = torch.from_numpy(obs_vec).float().unsqueeze(0).to(device)
                with torch.no_grad():
                    dist = policy(obs_tensor)
                    action = dist.sample().item()
                obs, reward, terminated, truncated, info = env.step(action)
                done = terminated or truncated
                episode_obs.append(obs_vec)
                episode_actions.append(np.array(action))
                episode_rewards.append(np.array(reward, dtype=np.float32))
            batch.append(episode_obs, episode_actions, episode_rewards, [ep]*len(episode_obs))
        train_batches.append(batch)

    # Compute advantages 
    for batch in train_batches:
        batch.compute_advantages(baseline, gae_lambda=1.0, normalize=True)
    
    # Compute gradients and adapt policy parameters
    params = None
    for batch in train_batches:
        loss = reinforce_loss(policy, batch, params=params)
        params = policy.update_params(loss, params=params, step_size=fast_lr, first_order=True)
    return params



def evaluate_policy(env, policy, preprocess_obs=None, params=None, render=False):
    with silence_sampling_rejected():
        obs, info = env.reset()
    steps = 0
    done = False
    success = False
    env_max_steps = getattr(env.unwrapped, 'max_steps', float('inf'))
    while not done and steps < env_max_steps:
        if render:
            env.render("human")
        obs_vec = preprocess_obs(obs)
        obs_tensor = torch.from_numpy(obs_vec).float().unsqueeze(0).to(device)
        with torch.no_grad():
            if params is not None:
                dist = policy(obs_tensor, params=params)
            else:
                dist = policy(obs_tensor)
            action = dist.sample().item()
        obs, reward, terminated, truncated, info = env.step(action)
        done = terminated or truncated
        steps += 1
        if terminated:
            success = True
    return steps, success


# Evaluation
N_MISSIONS = 10
N_EPISODES = 10

results_lang = []
results_maml = []
results_lang_conditioned = []
results_anil = []

success_lang = []
success_maml = []
success_lang_conditioned = []
success_anil = []

print("Comparing LA-MAML with baselines:")
for i in range(N_MISSIONS):
    mission = random.choice(missions)

    # 1. LA-MAML policy
    theta_prime = get_language_adapted_params(policy_lang, mission, mission_encoder, mission_adapter, device)
    lang_steps, lang_successes = [], []
    for ep in range(N_EPISODES):
        env.reset_task(mission)
        steps, success = evaluate_policy(env, policy_lang, preprocess_obs= sampler_lang.preprocess_obs, params=theta_prime)
        lang_steps.append(steps)
        lang_successes.append(success)
    results_lang.append(np.mean(lang_steps))
    success_lang.append(np.mean(lang_successes))

    # 2. MAML policy
    maml_params = adapt_policy_for_task(mission, policy_maml, num_steps=2, batch_size=10, baseline=baseline)
    maml_steps, maml_successes = [], []
    for ep in range(N_EPISODES):
        env.reset_task(mission)
        steps, success = evaluate_policy(env, policy_maml, preprocess_obs= sampler_maml.preprocess_obs, params=maml_params)
        maml_steps.append(steps)
        maml_successes.append(success)
    results_maml.append(np.mean(maml_steps))
    success_maml.append(np.mean(maml_successes))

    # 3. Language-conditioned policy
    ablation_steps, ablation_successes = [], []
    preprocess_ablation = lambda obs, m=mission: sampler_lang_policy.preprocess_obs(obs, mission_str=m)
    for ep in range(N_EPISODES):
        env.reset_task(mission)
        steps, success = evaluate_policy(env, policy_ablation, preprocess_obs=preprocess_ablation)
        ablation_steps.append(steps)
        ablation_successes.append(success)
    results_lang_conditioned.append(np.mean(ablation_steps))
    success_lang_conditioned.append(np.mean(ablation_successes))

    # 4. ANIL policy
    anil_params = adapt_anil_policy_for_task(mission, policy_anil, env, baseline, num_steps=2, batch_size=10, head_layer_prefix=head_layer_prefix)
    anil_steps, anil_successes = [], []
    for ep in range(N_EPISODES):
        env.reset_task(mission)
        steps, success = evaluate_policy(env, policy_anil, preprocess_obs=sampler_anil.preprocess_obs, params=anil_params)
        anil_steps.append(steps)
        anil_successes.append(success)
    results_anil.append(np.mean(anil_steps))
    success_anil.append(np.mean(anil_successes))

# Results
print("\n" + "="*50)
print(f"{'Policy':<25} | {'Avg Steps':<20}")
print("-" * 50)
print(f"{'LA-MAML':<25} | {np.mean(results_lang):<8.2f} ± {np.std(results_lang):<4.2f}")
print(f"{'MAML':<25} | {np.mean(results_maml):<8.2f} ± {np.std(results_maml):<4.2f}")
print(f"{'Language-conditioned':<25} | {np.mean(results_lang_conditioned):<8.2f} ± {np.std(results_lang_conditioned):<4.2f}")
print(f"{'ANIL':<25} | {np.mean(results_anil):<8.2f} ± {np.std(results_anil):<4.2f}")
print("="*50)

print("\n" + "="*40)
print(f"{'Policy':<25} | {'Success Rate':<10}")
print("-" * 40)
print(f"{'LA-MAML':<25} | {np.mean(success_lang)*100:.1f}%")
print(f"{'MAML':<25} | {np.mean(success_maml)*100:.1f}%")
print(f"{'Language-conditioned':<25} | {np.mean(success_lang_conditioned)*100:.1f}%")
print(f"{'ANIL':<25} | {np.mean(success_anil)*100:.1f}%")
print("="*40)

# Save to Excel
try:
    xlsx_path = "evaluation_results.xlsx"
    
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]
        
    if env_name in wb.sheetnames:
        ws = wb[env_name]
    else:
        ws = wb.create_sheet(env_name)
        # Add Headers
        ws.append(["Room Size", "Num Distractors", "Max Steps", "Delta Theta", 
                   "Avg Steps LA-MAML", "Success Prob LA-MAML", 
                   "Avg Steps MAML", "Success Prob MAML", 
                   "Avg Steps Lang-Cond", "Success Prob Lang-Cond", 
                   "Avg Steps ANIL", "Success Prob ANIL"])

    # Append Results
    ws.append([room_size, num_dists, max_steps, delta_theta, 
               f"{np.mean(results_lang):.2f} ± {np.std(results_lang):.2f}", np.mean(success_lang), 
               f"{np.mean(results_maml):.2f} ± {np.std(results_maml):.2f}", np.mean(success_maml), 
               f"{np.mean(results_lang_conditioned):.2f} ± {np.std(results_lang_conditioned):.2f}", np.mean(success_lang_conditioned), 
               f"{np.mean(results_anil):.2f} ± {np.std(results_anil):.2f}", np.mean(success_anil)])
    
    wb.save(xlsx_path)
    print(f"Results successfully logged to '{xlsx_path}' under sheet '{env_name}'")
except Exception as e:
    print(f"Failed to save to Excel: {e}")
