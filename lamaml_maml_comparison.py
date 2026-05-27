import torch
import numpy as np
import random
from functools import partial
from environment import (GoToLocalMissionEnv,
                         GoToOpenMissionEnv, 
                         GoToObjDoorMissionEnv,  
                         PickupDistMissionEnv,
                         OpenDoorMissionEnv, 
                         OpenDoorLocMissionEnv,
                         OpenDoorsOrderMissionEnv)
from sampler_lang import BabyAIMissionTaskWrapper, SentenceMissionEncoder, MissionParamAdapter
import sampler_lang 
import sampler_maml
import os
from maml_rl.policies.categorical_mlp import CategoricalMLPPolicy
from maml_rl.utils.reinforcement_learning import reinforce_loss
from maml_rl.episode import BatchEpisodes
from maml_rl.baseline import LinearFeatureBaseline
import argparse
from openpyxl import Workbook, load_workbook
import builtins, io
from contextlib import contextmanager, redirect_stdout, redirect_stderr

@contextmanager
def silence_sampling_rejected():
    real_print = builtins.print
    buf = io.StringIO()
    def filtered_print(*args, **kwargs):
        if args and isinstance(args[0], str) and args[0].startswith("Sampling rejected: unreachable object"):
            return
        return real_print(*args, **kwargs)
    builtins.print = filtered_print
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            yield
    finally:
        builtins.print = real_print


seed = 42
random.seed(seed)
np.random.seed(seed)
torch.manual_seed(seed)
if torch.cuda.is_available():
    torch.cuda.manual_seed_all(seed)

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

p = argparse.ArgumentParser()
p.add_argument("--env", dest="env_name",
               choices=["GoToLocal","PickupDist","GoToObjDoor","GoToOpen","OpenDoor",
                        "OpenDoorLoc","OpenDoorsOrder"],
               default="GoToLocal")
p.add_argument("--room-size", type=int, default=7)
p.add_argument("--num-dists", type=int, default=3)
p.add_argument("--max-steps", type=int, default=300)
p.add_argument("--delta-theta", type=float, default=0.3)
p.add_argument("--skip-random", action="store_true",
               help="Skip the random-policy baseline to speed up evaluation")

args = p.parse_args()


OBJECTS = ['box']
COLORS = ['red', 'green', 'blue', 'purple','yellow', 'grey']
PREP_LOCS = ['on', 'at', 'to']

# Location names
LOC_NAMES = ['right', 'front']

DOOR_COLORS = ['yellow', 'grey']

# For Pickup
PICKUP_MISSIONS = [f"pick up the {color} {obj}" for color in COLORS for obj in OBJECTS]

# For GoToLocal
LOCAL_MISSIONS = [f"go to the {color} {obj}" for color in COLORS for obj in OBJECTS]

# For environments that include doors (GoToObjDoor, GoToOpen, Open)
DOOR_MISSIONS = [f"go to the {color} door" for color in DOOR_COLORS]
OPEN_DOOR_MISSIONS = [f"open the {color} door" for color in DOOR_COLORS]
DOOR_LOC_MISSIONS = [f"open the door {prep} the {loc}" for prep in PREP_LOCS for loc in LOC_NAMES]
OPEN_DOORS_ORDER_MISSIONS = (
    [f"open the {c1} door" for c1 in DOOR_COLORS] +
    [f"open the {c1} door, then open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS] +
    [f"open the {c1} door after you open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS]
)


def build_env(env, room_size, num_dists, max_steps, missions):

    if env == "GoToLocal":
        base = GoToLocalMissionEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps)
    elif env == "PickupDist":
        base = PickupDistMissionEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps)
    elif env == "GoToObjDoor":
        base = GoToObjDoorMissionEnv(max_steps=max_steps, num_distractors=num_dists)
    elif env == "GoToOpen":
        base = GoToOpenMissionEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps)
    elif env == "OpenDoor":
        base = OpenDoorMissionEnv(room_size=room_size, max_steps=max_steps)
    elif env == "OpenDoorLoc":
        base = OpenDoorLocMissionEnv(room_size=room_size, max_steps=max_steps)
    elif env == "OpenDoorsOrder":
        base = OpenDoorsOrderMissionEnv(room_size=room_size)
    else:
        raise ValueError(f"Unknown env_name: {env}")

    return BabyAIMissionTaskWrapper(base, missions=missions)


def select_missions(env_name):
    mission_map = {
        "GoToLocal": LOCAL_MISSIONS,
        "PickupDist": PICKUP_MISSIONS,
        "GoToObjDoor": LOCAL_MISSIONS + DOOR_MISSIONS,
        "GoToOpen": LOCAL_MISSIONS,
        "OpenDoor": OPEN_DOOR_MISSIONS,
        "OpenDoorLoc": OPEN_DOOR_MISSIONS + DOOR_LOC_MISSIONS,
        "OpenDoorsOrder": OPEN_DOORS_ORDER_MISSIONS    }
    return mission_map[env_name]


env_name  = args.env_name
room_size = "env" if args.env_name in ["GoToObjDoor"] else args.room_size
num_dists = "env" if args.env_name in ["OpenDoor", "OpenDoorLoc", "OpenDoorsOrder"] else args.num_dists
max_steps = args.max_steps
delta_theta = args.delta_theta

missions = select_missions(env_name)
make_env = partial(build_env, env_name, room_size, num_dists, max_steps, missions)
env = make_env()

print(f"env name {env} \n")
print(f"room_size: {room_size}\nnum_dists: {num_dists}\nmax_steps: {max_steps}\n")


# ==================== SETUP: LA-MAML Policy ====================
ckpt = torch.load(f"lang_model/lang_{env_name}_{delta_theta}.pth", map_location=device)

dummy_obs, _ = env.reset()
input_size_lang = sampler_lang.preprocess_obs(dummy_obs).shape[0]
output_size = env.action_space.n
hidden_sizes = (64, 64)
nonlinearity = torch.nn.functional.tanh

# Policy language
policy_lang = CategoricalMLPPolicy(
    input_size=input_size_lang,
    output_size=output_size,
    hidden_sizes=hidden_sizes,
    nonlinearity=nonlinearity,
).to(device)  
policy_lang.load_state_dict(ckpt["policy"])
policy_lang.eval()
policy_param_shapes = [p.shape for p in policy_lang.parameters()]

mission_encoder = SentenceMissionEncoder(
    model_name="all-MiniLM-L6-v2",
    frozen=True,
    normalize=True,
    cache=True,
    device=device
)
mission_encoder.eval()

preprocess_obs = sampler_lang.preprocess_obs
mission_encoder_output_dim = mission_encoder.output_dim

# Adapter
mission_adapter = MissionParamAdapter(mission_encoder_output_dim, policy_param_shapes).to(device)
mission_adapter.load_state_dict(ckpt["mission_adapter"])    
mission_adapter.eval()


# ==================== SETUP: MAML Policy ====================
# restore saved maml policies (2-step and 3-step)
ckpt_base_2 = f"maml_model/maml_{env_name}_2"
ckpt_base_3 = f"maml_model/maml_{env_name}_3"

dummy_obs, _ = env.reset()
input_size_maml = sampler_maml.preprocess_obs(dummy_obs).shape[0]

# Policy MAML 2-step
policy_maml_2 = CategoricalMLPPolicy(
        input_size=input_size_maml,
        output_size=output_size,
        hidden_sizes=hidden_sizes,
        nonlinearity=nonlinearity,      
    ).to(device)

policy_maml_2.load_state_dict(torch.load(ckpt_base_2 + ".pth", map_location=device))
policy_maml_2.eval()

# Policy MAML 3-step
policy_maml_3 = CategoricalMLPPolicy(
        input_size=input_size_maml,
        output_size=output_size,
        hidden_sizes=hidden_sizes,
        nonlinearity=nonlinearity,      
    ).to(device)

policy_maml_3.load_state_dict(torch.load(ckpt_base_3 + ".pth", map_location=device))
policy_maml_3.eval()

baseline = LinearFeatureBaseline(input_size_maml).to(device)


def get_language_adapted_params(policy, mission_str, mission_encoder, mission_adapter, device):
    with torch.no_grad():
        mission_emb = mission_encoder(mission_str).to(device)
        delta_thetas = mission_adapter(mission_emb)
        delta_thetas = [delta * delta_theta for delta in delta_thetas]
    policy_params = list(policy.parameters())
    param_names = list(dict(policy.named_parameters()).keys())
    from collections import OrderedDict
    theta_prime = OrderedDict(
        (name, param + delta.squeeze(0))
        for name, param, delta in zip(param_names, policy_params, delta_thetas)
    )
    return theta_prime



def evaluate_policy(env, policy, preprocess_obs=None, params=None, max_steps=max_steps, render=False):
    with silence_sampling_rejected():
        obs, info = env.reset()
    steps = 0
    done = False
    success = False
    env_max_steps = getattr(env.unwrapped, 'max_steps', float('inf'))
    if max_steps == float('inf'):
        max_steps = env_max_steps
    while not done and steps < max_steps:
        if render:
            env.render("human")
        obs_vec = preprocess_obs(obs)
        obs_tensor = torch.from_numpy(obs_vec).float().unsqueeze(0).to(device)
        with torch.no_grad():
            if params is not None:
                dist = policy(obs_tensor, params=params)
            else:
                dist = policy(obs_tensor)
            action = dist.sample().item()
        obs, reward, terminated, truncated, info = env.step(action)
        done = terminated or truncated
        steps += 1
        if terminated:
            success = True
    return steps, success


# Evaluation
N_MISSIONS = 10
N_EPISODES = 10

results_lamaml = []
results_maml_2 = []   # MAML with 2-step inner loop
results_maml_3 = []   # MAML with 3-step inner loop

success_lamaml = []
success_maml_2 = []
success_maml_3 = []

print("Comparing LA-MAML policy with MAML 2-step and 3-step on random missions:")
for i in range(N_MISSIONS):
    mission = random.choice(missions)

    # 1. LA-MAML policy (language adaptation)
    theta_prime_lamaml = get_language_adapted_params(policy_lang, mission, mission_encoder, mission_adapter, device)
    lamaml_steps, lamaml_successes = [], []

    for ep in range(N_EPISODES):
        env.reset_task(mission)
        steps, success = evaluate_policy(env, policy_lang, preprocess_obs=sampler_lang.preprocess_obs, params=theta_prime_lamaml)
        lamaml_steps.append(steps)
        lamaml_successes.append(success)
    mean_lamaml = np.mean(lamaml_steps)
    results_lamaml.append(mean_lamaml)
    success_lamaml.append(np.mean(lamaml_successes))

    # 2. MAML 2-step
    maml_2_steps, maml_2_successes = [], []
    for ep in range(N_EPISODES):
        env.reset_task(mission)
        steps, success = evaluate_policy(
            env, policy_maml_2,
            preprocess_obs=sampler_maml.preprocess_obs,
            params=None
        )
        maml_2_steps.append(steps)
        maml_2_successes.append(success)
    mean_maml_2 = np.mean(maml_2_steps)
    results_maml_2.append(mean_maml_2)
    success_maml_2.append(np.mean(maml_2_successes))

    # 3. MAML 3-step
    maml_3_steps, maml_3_successes = [], []
    for ep in range(N_EPISODES):
        env.reset_task(mission)
        steps, success = evaluate_policy(
            env, policy_maml_3,
            preprocess_obs=sampler_maml.preprocess_obs,
            params=None
        )
        maml_3_steps.append(steps)
        maml_3_successes.append(success)
    mean_maml_3 = np.mean(maml_3_steps)
    results_maml_3.append(mean_maml_3)
    success_maml_3.append(np.mean(maml_3_successes))

# Results
print("\n" + "="*70)
print(f"{'Policy':<25} | {'Avg Steps':<20}")
print("-" * 70)
print(f"{'LA-MAML policy':<25} | {np.mean(results_lamaml):<8.2f} ± {np.std(results_lamaml):<4.2f}")
print(f"{'MAML 2-step':<25} | {np.mean(results_maml_2):<8.2f} ± {np.std(results_maml_2):<4.2f}")
print(f"{'MAML 3-step':<25} | {np.mean(results_maml_3):<8.2f} ± {np.std(results_maml_3):<4.2f}")
print("="*70)

print("\n" + "="*50)
print(f"{'Policy':<25} | {'Success Rate':<10}")
print("-" * 50)
print(f"{'LA-MAML policy':<25} | {np.mean(success_lamaml)*100:.1f}%")
print(f"{'MAML 2-step':<25} | {np.mean(success_maml_2)*100:.1f}%")
print(f"{'MAML 3-step':<25} | {np.mean(success_maml_3)*100:.1f}%")
print("="*50)

# Save to Excel
try:
    xlsx_path = "lamaml_maml_comparison_results.xlsx"
    
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]
        
    if env_name in wb.sheetnames:
        ws = wb[env_name]
    else:
        ws = wb.create_sheet(env_name)
        # Add Headers
        ws.append(["Room Size", "Num Distractors", "Max Steps", "Delta Theta", 
                   "Avg Steps LA-MAML", "Success Prob LA-MAML", 
                   "Avg Steps MAML 2-step", "Success Prob MAML 2-step", 
                   "Avg Steps MAML 3-step", "Success Prob MAML 3-step"])

    # Append Results
    ws.append([room_size, num_dists, max_steps, delta_theta, 
               f"{np.mean(results_lamaml):.2f} ± {np.std(results_lamaml):.2f}", np.mean(success_lamaml), 
               f"{np.mean(results_maml_2):.2f} ± {np.std(results_maml_2):.2f}", np.mean(success_maml_2), 
               f"{np.mean(results_maml_3):.2f} ± {np.std(results_maml_3):.2f}", np.mean(success_maml_3)])
    
    wb.save(xlsx_path)
    print(f"\nResults successfully logged to '{xlsx_path}' under sheet '{env_name}'")
except Exception as e:
    print(f"Failed to save to Excel: {e}")
